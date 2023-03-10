import openpyxl

path = "C:\\Users\\graveendran\\PycharmProjects\\pyTestProject\\dataFiles\\excelData.xlsx"
workbook = openpyxl.load_workbook(path)


def readLoginData():
    list1 = []
    sheet = workbook.get_sheet_by_name("Sheet1")
    rows = sheet.max_row

    for r in range(2, rows + 1):
        username = sheet.cell(r, 1).value
        password = sheet.cell(r, 2).value
        tuple1 = (username, password)
        print(tuple1)
        list1.append(tuple1)
    print(list1)
    return list1


def readSignInData():
    list1 = []
    sheet = workbook.get_sheet_by_name("Sheet2")
    rows = sheet.max_row

    for r in range(2, rows + 1):
        username = sheet.cell(r, 1).value
        password = sheet.cell(r, 2).value
        tuple1 = (username, password)
        print(tuple1)
        list1.append(tuple1)
    print(list1)
    return list1


def readContactData():
    list1 = []
    sheet = workbook.get_sheet_by_name("Sheet3")
    contactEmail = sheet.cell(1, 1).value
    contactName = sheet.cell(1, 2).value
    contactMessage = sheet.cell(1, 3).value
    tuple1 = (contactEmail, contactName, contactMessage)
    print(tuple1)
    list1.append(tuple1)
    return list1


def readPurchaseData():
    list1 = []
    sheet = workbook.get_sheet_by_name("Sheet4")
    name = sheet.cell(1, 1).value
    country = sheet.cell(1, 2).value
    city = sheet.cell(1, 3).value
    card = sheet.cell(1, 4).value
    month = sheet.cell(1, 5).value
    year = sheet.cell(1, 6).value
    tuple1 = (name, country, city, card, month, year)
    print(tuple1)
    list1.append(tuple1)
    return list1
